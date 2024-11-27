"""
   pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Tworzenie nowego skoroszytu
wb = Workbook()
ws = wb.active
ws.title = "Przykład danych"

# Dodawanie danych do arkusza
ws['A1'] = "Imię"
ws['B1'] = "Nazwisko"
ws['C1'] = "Wiek"

# Formatowanie nagłówków
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD",
                          end_color="4F81BD", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col in ['A', 'B', 'C']:
    ws[f'{col}1'].font = header_font
    ws[f'{col}1'].fill = header_fill
    ws[f'{col}1'].alignment = header_alignment

# Dodawanie przykładowych danych
data = [
    ["Anna", "Kowalska", 29],
    ["Jan", "Nowak", 34],
    ["Maria", "Wiśniewska", 27],
    ["Tomasz", "Kowalczyk", 42]
]

for row in data:
    ws.append(row)

# Formatowanie danych (np. wyrównanie tekstu w komórkach)
for row in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Dodawanie obramowania do komórek
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):
    for cell in row:
        cell.border = thin_border

# Ustawienie szerokości kolumn
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 10

# Zapis pliku
file_path = "D:\\Przyklad_openpyxl.xlsx"
wb.save(file_path)

file_path
